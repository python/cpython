"""Test: et-xmlfile"""
import sys
sys.stdout.reconfigure(line_buffering=True)
try:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    assert ws["A1"].value == "hello"
    print("et-xmlfile: PASS")
except Exception as e:
    print(f"et-xmlfile: FAIL: {e}")
    sys.exit(1)
